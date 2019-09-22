# -*- coding: utf-8 -*-

import os
import sys
import csv
import textwrap
import argparse
from argparse import ArgumentParser, Namespace
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side

class MultipleCSVToExcel:
    TEMPLATE_PATH = os.path.dirname(os.path.abspath(__file__)) + "\\template"
    CSV_PATH = os.path.dirname(os.path.abspath(__file__)) + "\\csv"
    TEMPLATE_FILE = "template.xlsx"
    SIDE = Side(border_style = "thin", color = "000000")
    BORDER = Border(top = SIDE, bottom = SIDE, left = SIDE, right = SIDE)

    def __init__(self, args: Namespace) -> None:
        self.DEST_ROW_INDEX = args.row
        self.DEST_COLUMN_INDEX = args.column
        self.DELIMITER = args.delimiter 
        self.FONT = Font(name = args.font, size = args.size)
        self.NEW_FILE = args.file

    def __paste_data(self, worksheet: Workbook, data: list) -> None:
        for row_index, row_value in enumerate(data):
            for column_index in range(len(row_value)):
                cell = worksheet.cell(row = self.DEST_ROW_INDEX + row_index, \
                                        column = self.DEST_COLUMN_INDEX + column_index)
                cell.value = row_value[column_index]
                cell.font = self.FONT
                cell.border = self.BORDER

    def __open_csv_file(self, worksheet: Workbook, csv_file: list) -> None:
        with open(self.CSV_PATH + "\\" + csv_file, "r") as f:
            csv_data = csv.reader(f, delimiter=",")
            self.__paste_data(worksheet, csv_data)

    def __open_workbook(self) -> Workbook:
        try:
            return load_workbook(self.TEMPLATE_PATH + "\\" + self.TEMPLATE_FILE)
        except FileNotFoundError as err:
            print(err)
            sys.exit()

    def __get_csv_files(self) -> list:
        try:
            csv_files = os.listdir(self.CSV_PATH)
            if not csv_files:
                print("ERROR: Nothing csv files in csv direcoty. please store csv file.")
                sys.exit()
        except FileNotFoundError as err:
            print(err)
            sys.exit()
        return csv_files

    def copy_csv_to_excel(self) -> None:
        csv_files = self.__get_csv_files()
        wb = self.__open_workbook()
        ws = wb.active
        for csv_file in csv_files:
            copy_ws = wb.copy_worksheet(ws)
            copy_ws.title = csv_file.split(self.DELIMITER, 1)[0]
            self.__open_csv_file(copy_ws, csv_file)
        wb.save(self.NEW_FILE)


def parser() -> Namespace:
    usage = textwrap.dedent('''\
            python {} [--help] [--file <new file name> ] [--row <number] [--column <number>]
                                  [--delimiter <delimiter>] [--font <font style>] [--size <font size> ]
                            '''.format(__file__))
    description = textwrap.dedent('''\
                    Copy multipule csv files to excel.
                    Required template.xlsx in template directory, and csv files in csv directories.
                                  ''')
    argparser = ArgumentParser(usage = usage, description = description, \
                                formatter_class = argparse.RawDescriptionHelpFormatter)
    argparser.add_argument("-f", "--file", type = str, default = "new_file.xlsx", \
                            metavar = "<new file name>", \
                            help = "file name for destination workbook")
    argparser.add_argument("-r", "--row", type = int, default = 1, \
                            metavar = "<number>", \
                            help = "destination row number")
    argparser.add_argument("-c", "--column", type = int, default = 1, \
                            metavar = "<number>", \
                            help = "destination column number")
    argparser.add_argument("-d", "--delimiter", type = str, default = ".", \
                            metavar = "<delimiter>", \
                            help = "delimiter from csv file name to destination sheet name")
    argparser.add_argument("--font", type = str, default = "游ゴシック", \
                            metavar = "<font style>", \
                            help = "destination font style")
    argparser.add_argument("--size", type = int, default = 11, \
                            metavar = "<font size>", \
                            help = "destination font size")
    return argparser.parse_args()


def main() -> None:
    args = parser()
    mcte = MultipleCSVToExcel(args)
    mcte.copy_csv_to_excel()

if __name__ == "__main__":
    main()